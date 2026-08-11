"""Streamlit 公共导航与示例流程烟雾测试。"""

import json
import tomllib
from dataclasses import replace
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook
from streamlit.testing.v1 import AppTest

from src.config import APP_NAME, APP_VERSION
from src.ui_common import (
    PUBLIC_PRIVACY_NOTICE,
    RESEARCH_DISCLAIMER,
    SIDEBAR_PRIVACY_NOTICE,
)

SENSITIVE_WARNING_TERMS = (
    "请勿上传",
    "账号密码",
    "API密钥",
    "交易凭证",
    "个人敏感信息",
    "商业机密",
    "受限制数据",
)


def _load_app() -> AppTest:
    return AppTest.from_file("app.py", default_timeout=20).run()


def _open_page(app: AppTest, page_name: str) -> AppTest:
    app.radio(key="app_navigation").set_value(page_name).run()
    return app


def _download_labels(app: AppTest) -> list[str]:
    return [button.label for button in app.get("download_button")]


def _visible_text(app: AppTest) -> str:
    element_types = (
        "caption",
        "error",
        "info",
        "markdown",
        "success",
        "title",
        "warning",
        "text",
    )
    values: list[str] = []
    for element_type in element_types:
        values.extend(str(element.value) for element in app.get(element_type))
    return "\n".join(values)


def _multisheet_xlsx() -> bytes:
    """在内存中生成用于 AppTest 的两工作表 XLSX。"""
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "说明"
    first_sheet.append(["text"])
    first_sheet.append(["demo"])
    data_sheet = workbook.create_sheet("数据")
    data_sheet.append(["日期", "收益"])
    data_sheet.append(["2026-01-01", 0.01])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _mapping_multisheet_xlsx() -> bytes:
    """生成两份可分别确认收益率与净值主口径的工作表。"""
    workbook = Workbook()
    return_sheet = workbook.active
    return_sheet.title = "收益表"
    return_sheet.append(["trade_date", "strategy_return"])
    return_sheet.append(["2026-01-01", 0.01])
    return_sheet.append(["2026-01-02", -0.02])
    return_sheet.append(["2026-01-03", 0.03])
    return_sheet.append(["2026-01-04", 0.01])
    nav_sheet = workbook.create_sheet("净值表")
    nav_sheet.append(["trade_date", "strategy_nav"])
    nav_sheet.append(["2026-01-01", 1.0])
    nav_sheet.append(["2026-01-02", 0.98])
    nav_sheet.append(["2026-01-03", 1.01])
    nav_sheet.append(["2026-01-04", 1.02])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _open_general_csv(content: str) -> AppTest:
    app = _open_page(_load_app(), "单实验分析")
    app.radio(key="single_data_mode").set_value("通用文件导入（CSV/XLSX）").run()
    app.get("file_uploader")[0].upload(
        "mapping.csv",
        content.encode("utf-8"),
        "text/csv",
    ).run()
    return app


def _mapping_basis_selectbox(app: AppTest):
    return next(selectbox for selectbox in app.selectbox if selectbox.label == "策略分析主口径")


def _mapping_acknowledgement(app: AppTest):
    return next(
        checkbox for checkbox in app.checkbox if checkbox.label.startswith("我已核对字段含义")
    )


def _submit_mapping(app: AppTest) -> AppTest:
    _mapping_acknowledgement(app).set_value(True)
    next(button for button in app.button if button.label == "确认字段映射").click()
    return app.run()


def _generate_standardization_preview(app: AppTest) -> AppTest:
    next(button for button in app.button if button.label == "生成标准化预览").click()
    return app.run()


def _execute_strict_validation(app: AppTest) -> AppTest:
    next(button for button in app.button if button.label == "执行严格协议验证").click()
    return app.run()


def _generic_final_confirmation(app: AppTest):
    return next(
        checkbox
        for checkbox in app.checkbox
        if checkbox.label.startswith("我已核对日期、收益率或净值字段")
    )


def _start_generic_analysis(app: AppTest) -> AppTest:
    _generic_final_confirmation(app).set_value(True)
    next(button for button in app.button if button.label == "开始绩效分析").click()
    return app.run()


def _strict_protocol_summary(app: AppTest):
    return next(
        item.value.set_index("项目")["结果"]
        for item in app.dataframe
        if item.value.columns.tolist() == ["项目", "结果"]
        and "严格协议" in item.value["项目"].tolist()
    )


def _ready_generic_return(*, benchmark: bool = False) -> AppTest:
    benchmark_header = ",benchmark_return" if benchmark else ""
    benchmark_values = [",0.004", ",-0.002", ",0.006", ",0.001"]
    rows = []
    for index, (date, value) in enumerate(
        (
            ("2026-01-01", "0.01"),
            ("2026-01-02", "-0.02"),
            ("2026-01-03", "0.03"),
            ("2026-01-04", "0.01"),
        )
    ):
        rows.append(f"{date},{value}{benchmark_values[index] if benchmark else ''}")
    app = _open_general_csv(
        f"trade_date,strategy_return{benchmark_header}\n" + "\n".join(rows) + "\n"
    )
    _submit_mapping(app)
    return _generate_standardization_preview(app)


def _ready_generic_nav() -> AppTest:
    app = _open_general_csv(
        "trade_date,strategy_nav,daily_ret\n"
        "2026-01-01,100.0,0.0\n"
        "2026-01-02,102.0,0.02\n"
        "2026-01-03,101.0,-0.00980392156862745\n"
        "2026-01-04,104.0,0.0297029702970297\n"
    )
    _set_mapping_and_submit(
        app,
        basis="策略净值为主",
        roles={
            "date": "trade_date",
            "strategy_nav": "strategy_nav",
            "daily_ret": "daily_ret",
        },
    )
    return _generate_standardization_preview(app)


def _mapping_role_selectbox(app: AppTest, role: str):
    return next(selectbox for selectbox in app.selectbox if selectbox.label.startswith(f"{role} ·"))


def _set_mapping_and_submit(
    app: AppTest,
    *,
    basis: str,
    roles: dict[str, str],
) -> AppTest:
    _mapping_basis_selectbox(app).set_value(basis)
    for role, column_name in roles.items():
        _mapping_role_selectbox(app, role).set_value(column_name)
    return _submit_mapping(app)


def test_home_page_is_default_and_has_no_analysis_charts() -> None:
    app = _load_app()

    assert not app.exception
    assert app.radio(key="app_navigation").value == "首页"
    assert app.title[0].value == APP_NAME
    assert len(app.get("plotly_chart")) == 0


def test_home_page_explains_product_and_main_workflows() -> None:
    app = _load_app()
    page_text = _visible_text(app)

    assert APP_NAME in page_text
    assert "量化研究实验台" in page_text
    assert "字段核验、绩效分析、可视化与标准化导出" in page_text
    assert "单实验分析" in page_text
    assert "多实验比较" in page_text
    assert "1. 上传数据" in page_text
    assert "4. 下载结果" in page_text


@pytest.mark.parametrize(
    ("button_key", "expected_page"),
    (
        ("home_open_single", "单实验分析"),
        ("home_open_comparison", "多实验比较"),
        ("home_open_reference_files", "参考文件"),
    ),
)
def test_home_entry_buttons_open_expected_page(
    button_key: str,
    expected_page: str,
) -> None:
    app = _load_app()

    app.button(key=button_key).click().run()

    assert not app.exception
    assert app.radio(key="app_navigation").value == expected_page
    assert app.title[0].value == expected_page


def test_public_streamlit_configuration_is_valid() -> None:
    config_path = Path(".streamlit/config.toml")
    with config_path.open("rb") as config_file:
        config = tomllib.load(config_file)

    assert config["client"]["toolbarMode"] == "minimal"
    assert config["theme"]["base"] == "light"
    assert config["theme"]["primaryColor"] == "#0F766E"


def test_can_enter_single_analysis_and_sample_still_renders() -> None:
    app = _open_page(_load_app(), "单实验分析")

    assert not app.exception
    assert app.title[0].value == "单实验分析"
    assert app.radio(key="single_data_mode").value == "使用示例数据"
    assert len(app.get("plotly_chart")) == 2
    assert "下载标准日频收益 CSV 模板" in _download_labels(app)


def test_single_page_exposes_strict_and_general_import_paths() -> None:
    app = _open_page(_load_app(), "单实验分析")
    data_mode = app.radio(key="single_data_mode")

    assert data_mode.options == [
        "使用示例数据",
        "按现有标准协议上传",
        "通用文件导入（CSV/XLSX）",
    ]


def test_general_import_waiting_state_does_not_render_performance_results() -> None:
    app = _open_page(_load_app(), "单实验分析")
    app.radio(key="single_data_mode").set_value("通用文件导入（CSV/XLSX）").run()
    page_text = _visible_text(app)

    assert not app.exception
    assert "上传 1 份 CSV 或 XLSX 文件" in [uploader.label for uploader in app.get("file_uploader")]
    assert "当前尚未进行字段映射确认或绩效计算" in page_text
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0


def test_general_csv_upload_renders_preview_without_performance() -> None:
    app = _open_page(_load_app(), "单实验分析")
    app.radio(key="single_data_mode").set_value("通用文件导入（CSV/XLSX）").run()
    app.get("file_uploader")[0].upload(
        "preview.csv",
        "日期;收益\n2026-01-01;0.01\n".encode(),
        "text/csv",
    ).run()
    page_text = _visible_text(app)

    assert not app.exception
    assert "当前编码：" in page_text
    assert "当前分隔符：" in page_text
    assert any("文件已成功读取" in item.value for item in app.success)
    assert "字段识别建议" in page_text
    assert (
        "字段识别结果仅为确定性规则生成的建议。系统尚未建立字段映射，也不会使用当前文件计算绩效。"
    ) in page_text
    assert (
        "当前结果仅用于帮助确认字段。系统尚未建立字段映射，"
        "不会基于这些建议计算收益、净值、回撤或其他绩效指标。"
    ) in page_text
    assert "确认字段映射" in page_text
    assert _mapping_basis_selectbox(app).options == [
        "请选择",
        "策略收益率为主",
        "策略净值为主",
    ]
    assert _mapping_acknowledgement(app).value is False
    assert not any(item.value == "字段映射已确认。" for item in app.success)
    assert "date（日期）" in app.dataframe[1].value["业务角色"].tolist()
    assert len(app.dataframe) == 4
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0


def test_general_mapping_missing_required_fields_is_blocked() -> None:
    app = _open_general_csv("value,notes\n1,a\n2,b\n3,c\n")

    next(button for button in app.button if button.label == "确认字段映射").click().run()

    assert not app.exception
    errors = [item.value for item in app.error]
    assert "请选择策略分析主口径。" in errors
    assert "必须映射日期字段 date。" in errors
    assert "字段映射存在阻断性错误，尚未建立确认状态。" in errors
    assert not any(item.value == "字段映射已确认。" for item in app.success)


def test_general_return_mapping_can_be_explicitly_confirmed_without_analysis() -> None:
    app = _open_general_csv(
        "trade_date,strategy_return\n"
        "2026-01-01,0.01\n"
        "2026-01-02,-0.02\n"
        "2026-01-03,0.03\n"
        "2026-01-04,0.01\n"
    )

    assert _mapping_basis_selectbox(app).value == "策略收益率为主"
    assert not any(item.value == "字段映射已确认。" for item in app.success)

    _submit_mapping(app)
    page_text = _visible_text(app)

    assert not app.exception
    assert any(item.value == "字段映射已确认。" for item in app.success)
    assert "已确认映射摘要" in page_text
    assert "本节不会自动转换收益率单位" in page_text
    assert "尚未执行严格协议验证" not in page_text
    summary = app.dataframe[-1].value
    return_row = summary.loc[summary["业务角色"] == "strategy_return（策略收益率）"].iloc[0]
    assert return_row["原始字段"] == "strategy_return"
    assert return_row["是否为主口径字段"] == "是"
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0


def test_general_nav_mapping_can_be_explicitly_confirmed_without_analysis() -> None:
    app = _open_general_csv(
        "trade_date,strategy_nav\n"
        "2026-01-01,1.00\n"
        "2026-01-02,0.98\n"
        "2026-01-03,1.01\n"
        "2026-01-04,1.02\n"
    )

    assert _mapping_basis_selectbox(app).value == "策略净值为主"
    _submit_mapping(app)

    assert not app.exception
    assert any(item.value == "字段映射已确认。" for item in app.success)
    summary = app.dataframe[-1].value
    nav_row = summary.loc[summary["业务角色"] == "strategy_nav（策略净值）"].iloc[0]
    assert nav_row["原始字段"] == "strategy_nav"
    assert nav_row["是否为主口径字段"] == "是"
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0


def test_standardization_button_requires_confirmed_mapping() -> None:
    app = _open_general_csv("trade_date,strategy_return\n2026-01-01,0.01\n2026-01-02,-0.02\n")

    assert not app.exception
    assert "生成标准化预览" not in [button.label for button in app.button]


def test_confirmed_mapping_shows_button_but_does_not_run_automatically() -> None:
    app = _open_general_csv(
        "trade_date,strategy_return\n2026-01-01,0.01\n2026-01-02,-0.02\n2026-01-03,0.03\n"
    )
    _submit_mapping(app)
    page_text = _visible_text(app)

    assert "生成标准化预览" in [button.label for button in app.button]
    assert "尚未生成预览" in page_text
    assert "预检摘要" not in page_text


def test_return_standardization_preview_is_read_only_and_has_no_analysis_outputs() -> None:
    app = _open_general_csv(
        "trade_date,strategy_return,benchmark_return,strategy_nav\n"
        "2026-01-01,0.01,0.005,1.00\n"
        "2026-01-02,-0.02,-0.01,0.98\n"
        "2026-01-03,0.03,0.02,1.0094\n"
        "2026-01-04,0.01,0.005,1.019494\n"
    )
    _set_mapping_and_submit(
        app,
        basis="策略收益率为主",
        roles={"date": "trade_date", "strategy_return": "strategy_return"},
    )
    _generate_standardization_preview(app)
    page_text = _visible_text(app)

    assert not app.exception
    assert "标准化转换与数据质量预检" in page_text
    assert "标准化分析候选表（前 20 行）" in page_text
    assert "当前尚未执行绩效计算、图表生成、报告生成或结果导出" in page_text
    summary = next(
        item.value for item in app.dataframe if item.value.columns.tolist() == ["项目", "结果"]
    ).set_index("项目")["结果"]
    assert summary["标准化结构类型"] == "收益率分析候选表"
    assert summary["分析候选字段"] == "date、strategy_return、benchmark_return"
    assert "strategy_nav" in summary["诊断字段"]
    candidate_frames = [
        item.value
        for item in app.dataframe
        if item.value.columns.tolist() == ["date", "strategy_return", "benchmark_return"]
    ]
    assert len(candidate_frames) == 1
    assert "strategy_nav" not in candidate_frames[0]
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0
    assert "累计收益率" not in page_text
    assert "净值图" not in page_text
    assert "回撤图" not in page_text


def test_nav_standardization_preview_uses_nav_candidate_structure() -> None:
    app = _open_general_csv(
        "trade_date,strategy_nav,daily_ret,benchmark_nav\n"
        "2026-01-01,1.00,0.00,100.0\n"
        "2026-01-02,0.98,-0.02,99.0\n"
        "2026-01-03,1.0094,0.03,101.0\n"
        "2026-01-04,1.019494,0.01,102.0\n"
    )
    _set_mapping_and_submit(
        app,
        basis="策略净值为主",
        roles={
            "date": "trade_date",
            "strategy_nav": "strategy_nav",
            "daily_ret": "daily_ret",
        },
    )
    _generate_standardization_preview(app)

    assert not app.exception
    candidate_frames = [
        item.value
        for item in app.dataframe
        if item.value.columns.tolist() == ["date", "nav_strat", "daily_ret"]
    ]
    assert len(candidate_frames) == 1
    assert "strategy_return" not in candidate_frames[0]
    assert all("benchmark_return" not in item.value.columns for item in app.dataframe)


def test_standardization_blocking_issues_are_visible_without_exception() -> None:
    app = _open_general_csv(
        "trade_date,strategy_return\n2026-01-01,0.01\n01/02/2026,1.2%\n2026-01-03,-1.0\n"
    )
    _set_mapping_and_submit(
        app,
        basis="策略收益率为主",
        roles={"date": "trade_date", "strategy_return": "strategy_return"},
    )
    _generate_standardization_preview(app)
    page_text = _visible_text(app)

    assert not app.exception
    assert "标准化预检未通过" in page_text
    issue_frames = [item.value for item in app.dataframe if "问题代码" in item.value.columns]
    assert len(issue_frames) == 1
    assert {"date_unparseable", "numeric_unparseable", "return_at_or_below_minus_one"} <= set(
        issue_frames[0]["问题代码"]
    )


def test_standardization_warnings_are_visible_without_claiming_analysis_complete() -> None:
    app = _open_general_csv(
        "trade_date,strategy_return\n"
        "2026-01-01,0.01\n"
        "2026-01-02,0.02\n"
        "2026-01-03,0.03\n"
        "2026-01-04,0.04\n"
    )
    _submit_mapping(app)
    _generate_standardization_preview(app)
    page_text = _visible_text(app)

    assert not app.exception
    summary = next(
        item.value for item in app.dataframe if item.value.columns.tolist() == ["项目", "结果"]
    )
    assert "warning数量" in summary["项目"].tolist()
    assert "标准化预检通过，可以在下一阶段进入现有严格协议验证" in page_text
    assert "分析已经完成" not in page_text
    assert "数据验证全部通过" not in page_text


def test_failed_b4a_preview_does_not_offer_strict_protocol_button() -> None:
    app = _open_general_csv("trade_date,strategy_return\n2026-01-01,0.01\n2026-01-02,-1.0\n")
    _submit_mapping(app)
    _generate_standardization_preview(app)

    assert "标准化预检未通过" in _visible_text(app)
    assert "执行严格协议验证" not in [button.label for button in app.button]
    assert len(app.get("metric")) == 0


def test_valid_preview_exposes_strict_gate_without_running_analysis() -> None:
    app = _ready_generic_return()

    assert "执行严格协议验证" in [button.label for button in app.button]
    assert "标准化预检通过不等于现有严格协议验证通过" in _visible_text(app)
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0


def test_strict_validation_succeeds_but_still_requires_final_confirmation() -> None:
    app = _execute_strict_validation(_ready_generic_return())
    page_text = _visible_text(app)

    assert not app.exception
    assert "严格协议验证摘要" in page_text
    assert _strict_protocol_summary(app)["严格协议"] == "标准日收益协议"
    assert "现有严格协议验证通过" in page_text
    assert _generic_final_confirmation(app).value is False
    start_button = next(button for button in app.button if button.label == "开始绩效分析")
    assert start_button.disabled
    assert len(app.get("metric")) == 0
    assert len(app.get("download_button")) == 0


def test_return_primary_final_confirmation_reuses_full_analysis_output() -> None:
    app = _start_generic_analysis(_execute_strict_validation(_ready_generic_return()))
    page_text = _visible_text(app)

    assert not app.exception
    assert "数据来源：通用文件导入 · 用户确认映射" in page_text
    assert "核心指标" in page_text
    assert len(app.get("metric")) == 8
    assert len(app.get("plotly_chart")) == 2
    assert _download_labels(app) == ["下载分析报告", "下载标准化分析数据"]


def test_return_primary_with_benchmark_reuses_existing_benchmark_output() -> None:
    app = _start_generic_analysis(_execute_strict_validation(_ready_generic_return(benchmark=True)))

    assert not app.exception
    assert any(metric.label == "基准累计收益" for metric in app.metric)
    assert len(app.get("plotly_chart")) == 2


def test_nav_primary_reuses_adapter_metrics_charts_report_and_downloads() -> None:
    app = _start_generic_analysis(_execute_strict_validation(_ready_generic_nav()))
    page_text = _visible_text(app)

    assert not app.exception
    assert _strict_protocol_summary(app)["严格协议"] == "净值适配协议"
    assert "净值适配器一致性摘要" in page_text
    assert any(metric.label == "净值观察日数" for metric in app.metric)
    assert any(metric.label == "有效收益日数" for metric in app.metric)
    assert len(app.get("plotly_chart")) == 2
    assert _download_labels(app) == ["下载分析报告", "下载标准化分析数据"]
    assert "基准累计收益" not in [metric.label for metric in app.metric]


def test_strict_protocol_failure_is_controlled_and_never_shows_analysis() -> None:
    app = _ready_generic_return()
    preview = app.session_state["qrw_standardization:result"]
    invalid_frame = preview.analysis_frame.assign(unsupported=1)
    app.session_state["qrw_standardization:result"] = replace(
        preview,
        analysis_frame=invalid_frame,
    )
    app.run()
    _execute_strict_validation(app)
    page_text = _visible_text(app)

    assert not app.exception
    assert "现有严格协议验证未通过" in page_text
    assert "当前阶段不支持的字段" in page_text
    assert "开始绩效分析" not in [button.label for button in app.button]
    assert len(app.get("metric")) == 0
    assert len(app.get("download_button")) == 0


def test_regenerating_standardization_invalidates_strict_and_analysis_results() -> None:
    app = _start_generic_analysis(_execute_strict_validation(_ready_generic_return()))
    assert len(app.get("metric")) == 8

    _generate_standardization_preview(app)
    page_text = _visible_text(app)

    assert not app.exception
    assert "请重新执行严格协议验证" in page_text
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0


def test_mapping_change_invalidates_generic_analysis_results() -> None:
    app = _start_generic_analysis(_execute_strict_validation(_ready_generic_return()))
    _mapping_role_selectbox(app, "strategy_return").set_value("不映射").run()

    assert not app.exception
    assert "字段映射选择已变化，请重新确认字段映射" in _visible_text(app)
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0


def test_file_change_invalidates_generic_analysis_results() -> None:
    app = _start_generic_analysis(_execute_strict_validation(_ready_generic_return()))
    app.get("file_uploader")[0].upload(
        "replacement.csv",
        (b"trade_date,strategy_return\n2026-02-01,0.02\n2026-02-02,-0.01\n2026-02-03,0.01\n"),
        "text/csv",
    ).run()

    assert not app.exception
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0
    assert "重新确认字段映射" in _visible_text(app)


def test_xlsx_sheet_change_invalidates_generic_analysis_results() -> None:
    app = _open_page(_load_app(), "单实验分析")
    app.radio(key="single_data_mode").set_value("通用文件导入（CSV/XLSX）").run()
    app.get("file_uploader")[0].upload(
        "mapping.xlsx",
        _mapping_multisheet_xlsx(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).run()
    app.selectbox(key="general_xlsx_sheet").set_value("收益表").run()
    _submit_mapping(app)
    _generate_standardization_preview(app)
    _execute_strict_validation(app)
    _start_generic_analysis(app)
    assert len(app.get("metric")) == 8

    app.selectbox(key="general_xlsx_sheet").set_value("净值表").run()

    assert not app.exception
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0
    assert "重新确认字段映射" in _visible_text(app)


def test_mapping_change_invalidates_old_standardization_preview() -> None:
    app = _open_general_csv(
        "trade_date,strategy_return,alt_return\n"
        "2026-01-01,0.01,0.02\n"
        "2026-01-02,-0.02,-0.01\n"
        "2026-01-03,0.03,0.01\n"
    )
    _submit_mapping(app)
    _generate_standardization_preview(app)
    assert "预检摘要" in _visible_text(app)

    _mapping_role_selectbox(app, "strategy_return").set_value("alt_return").run()

    assert not app.exception
    assert "文件、解析设置或字段映射已变化，请重新生成标准化预览" in _visible_text(app)
    assert "预检摘要" not in _visible_text(app)
    assert "字段映射选择已变化，请重新确认字段映射" in _visible_text(app)


def test_replacing_file_invalidates_old_standardization_preview() -> None:
    app = _open_general_csv(
        "trade_date,strategy_return\n2026-01-01,0.01\n2026-01-02,-0.02\n2026-01-03,0.03\n"
    )
    _submit_mapping(app)
    _generate_standardization_preview(app)

    app.get("file_uploader")[0].upload(
        "replacement.csv",
        (b"trade_date,strategy_nav\n2026-01-01,1.00\n2026-01-02,0.98\n2026-01-03,1.01\n"),
        "text/csv",
    ).run()

    assert not app.exception
    assert "文件、解析设置或字段映射已变化，请重新生成标准化预览" in _visible_text(app)
    assert "预检摘要" not in _visible_text(app)


def test_switching_xlsx_sheet_invalidates_old_standardization_preview() -> None:
    app = _open_page(_load_app(), "单实验分析")
    app.radio(key="single_data_mode").set_value("通用文件导入（CSV/XLSX）").run()
    app.get("file_uploader")[0].upload(
        "mapping.xlsx",
        _mapping_multisheet_xlsx(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).run()
    app.selectbox(key="general_xlsx_sheet").set_value("收益表").run()
    _submit_mapping(app)
    _generate_standardization_preview(app)

    app.selectbox(key="general_xlsx_sheet").set_value("净值表").run()

    assert not app.exception
    assert "文件、解析设置或字段映射已变化，请重新生成标准化预览" in _visible_text(app)
    assert "预检摘要" not in _visible_text(app)


def test_switching_xlsx_sheet_invalidates_confirmed_mapping() -> None:
    app = _open_page(_load_app(), "单实验分析")
    app.radio(key="single_data_mode").set_value("通用文件导入（CSV/XLSX）").run()
    app.get("file_uploader")[0].upload(
        "mapping.xlsx",
        _mapping_multisheet_xlsx(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).run()
    app.selectbox(key="general_xlsx_sheet").set_value("收益表").run()
    _submit_mapping(app)

    assert any(item.value == "字段映射已确认。" for item in app.success)

    app.selectbox(key="general_xlsx_sheet").set_value("净值表").run()

    assert not app.exception
    assert any(item.value == "文件或解析设置已变化，请重新确认字段映射。" for item in app.warning)
    assert not any(item.value == "字段映射已确认。" for item in app.success)
    assert _mapping_basis_selectbox(app).value == "策略净值为主"


def test_replacing_uploaded_file_invalidates_confirmed_mapping() -> None:
    app = _open_general_csv(
        "trade_date,strategy_return\n"
        "2026-01-01,0.01\n"
        "2026-01-02,-0.02\n"
        "2026-01-03,0.03\n"
        "2026-01-04,0.01\n"
    )
    _submit_mapping(app)
    assert any(item.value == "字段映射已确认。" for item in app.success)

    app.get("file_uploader")[0].upload(
        "replacement.csv",
        (
            b"trade_date,strategy_nav\n"
            b"2026-01-01,1.00\n"
            b"2026-01-02,0.98\n"
            b"2026-01-03,1.01\n"
            b"2026-01-04,1.02\n"
        ),
        "text/csv",
    ).run()

    assert not app.exception
    assert any(item.value == "文件或解析设置已变化，请重新确认字段映射。" for item in app.warning)
    assert not any(item.value == "字段映射已确认。" for item in app.success)


def test_general_xlsx_upload_can_switch_selected_sheet() -> None:
    app = _open_page(_load_app(), "单实验分析")
    app.radio(key="single_data_mode").set_value("通用文件导入（CSV/XLSX）").run()
    app.get("file_uploader")[0].upload(
        "preview.xlsx",
        _multisheet_xlsx(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ).run()

    assert not app.exception
    assert app.selectbox(key="general_xlsx_sheet").options == ["说明", "数据"]
    assert app.selectbox(key="general_xlsx_sheet").value is None
    assert len(app.dataframe) == 0

    app.selectbox(key="general_xlsx_sheet").set_value("数据").run()
    page_text = _visible_text(app)

    assert not app.exception
    assert "当前工作表：** 数据" in page_text
    assert "工作表数量：** 2" in page_text
    assert "字段识别建议" in page_text
    assert len(app.dataframe) == 4
    assert len(app.get("metric")) == 0
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("download_button")) == 0


def test_existing_strict_protocol_upload_path_remains_available() -> None:
    app = _open_page(_load_app(), "单实验分析")
    app.radio(key="single_data_mode").set_value("按现有标准协议上传").run()

    assert not app.exception
    assert app.radio(key="single_data_format").options == [
        "标准日频收益 CSV",
        "每周调仓净值 CSV",
    ]
    assert "上传 1 份 CSV 文件" in [uploader.label for uploader in app.get("file_uploader")]


def test_optional_experiment_information_is_collapsed_and_usable() -> None:
    app = _open_page(_load_app(), "单实验分析")
    experiment_expander = next(item for item in app.expander if item.label == "实验信息（可选）")

    assert experiment_expander.proto.expanded is False
    assert app.text_input(key="experiment_name:sample").value == "示例日频收益实验"

    app.text_input(key="experiment_name:sample").set_value("折叠区域实验")
    app.text_input(key="strategy_name:sample").set_value("示例策略")
    app.text_area(key="research_notes:sample").set_value("表单仍可正常填写。")
    app.run()

    assert not app.exception
    assert app.text_input(key="experiment_name:sample").value == "折叠区域实验"
    assert app.text_input(key="strategy_name:sample").value == "示例策略"
    assert app.text_area(key="research_notes:sample").value == "表单仍可正常填写。"


def test_can_enter_comparison_and_sample_still_renders() -> None:
    app = _open_page(_load_app(), "多实验比较")

    assert not app.exception
    assert app.title[0].value == "多实验比较"
    assert app.radio(key="comparison_source_mode").value == "使用比较示例数据"
    assert len(app.get("plotly_chart")) == 2
    assert "下载标准化比较 CSV 模板" in _download_labels(app)


def test_reference_files_page_renders_synthetic_library_without_analysis() -> None:
    app = _open_page(_load_app(), "参考文件")
    page_text = _visible_text(app)

    assert not app.exception
    assert app.title[0].value == "参考文件"
    assert "确定性合成数据" in page_text
    assert "不代表真实投资结果" in page_text
    assert "参考文件" in page_text
    assert len(app.get("plotly_chart")) == 0
    assert len(app.get("file_uploader")) == 0
    assert "核心指标" not in page_text


def test_reference_page_exposes_all_static_downloads_with_safe_labels() -> None:
    app = _open_page(_load_app(), "参考文件")
    download_labels = _download_labels(app)

    assert not app.exception
    assert len(download_labels) == 11
    assert sum(label.startswith("下载错误示例：") for label in download_labels) == 5
    assert "下载标准收益率（含基准）" in download_labels
    assert "下载中文通用收益率示例" in download_labels
    assert "下载中文通用净值示例" in download_labels
    assert "下载多工作表线上回归 XLSX" in download_labels
    assert "下载错误示例：百分号收益率" in download_labels
    assert "下载错误示例：重复日期" in download_labels


def test_reference_page_groups_errors_in_collapsed_warning_section() -> None:
    app = _open_page(_load_app(), "参考文件")
    expander_labels = [expander.label for expander in app.expander]
    page_text = _visible_text(app)

    assert not app.exception
    assert "错误示例与预检说明" in expander_labels
    error_expander = next(item for item in app.expander if item.label == "错误示例与预检说明")
    assert error_expander.proto.expanded is False
    assert "故意包含不明确或不安全的数据" in page_text
    assert "不应作为正常分析模板" in page_text


def test_reference_page_shows_entry_basis_mapping_and_expected_result() -> None:
    app = _open_page(_load_app(), "参考文件")
    page_text = _visible_text(app)

    assert "推荐入口" in page_text
    assert "按现有标准协议上传" in page_text
    assert "通用文件导入（CSV/XLSX）" in page_text
    assert "推荐主口径" in page_text
    assert "策略收益率为主" in page_text
    assert "策略净值为主" in page_text
    assert "推荐字段映射" in page_text
    assert "交易日期 → date" in page_text
    assert "预期结果" in page_text


def test_reference_navigation_is_fifth_page_in_expected_order() -> None:
    app = _load_app()

    assert app.radio(key="app_navigation").options == [
        "首页",
        "单实验分析",
        "多实验比较",
        "参考文件",
        "使用说明",
    ]


def test_comparison_drawdown_axis_uses_two_decimal_percentage() -> None:
    app = _open_page(_load_app(), "多实验比较")
    drawdown_spec = json.loads(app.get("plotly_chart")[1].proto.spec)

    assert not app.exception
    assert drawdown_spec["layout"]["yaxis"]["tickformat"] == ".2%"


def test_can_enter_usage_guide() -> None:
    app = _open_page(_load_app(), "使用说明")
    tab_labels = [tab.label for tab in app.tabs]

    assert not app.exception
    assert app.title[0].value == "使用说明"
    assert tab_labels == ["单实验", "多实验", "参考文件", "常见错误", "数据处理"]
    assert "0.01 表示 1%" in _visible_text(app)
    assert "错误示例用于理解系统的阻断机制" in _visible_text(app)
    assert "不会自动切换页面、建立映射或启动分析" in _visible_text(app)


def test_page_displays_version_privacy_notice_and_disclaimer() -> None:
    app = _load_app()
    page_text = _visible_text(app)

    assert APP_VERSION in page_text
    assert PUBLIC_PRIVACY_NOTICE in page_text
    assert RESEARCH_DISCLAIMER in page_text


def test_sidebar_uses_concise_privacy_notice() -> None:
    app = _load_app()
    sidebar_warnings = [warning.value for warning in app.sidebar.warning]

    assert sidebar_warnings == [SIDEBAR_PRIVACY_NOTICE]
    assert SIDEBAR_PRIVACY_NOTICE == (
        "公开云端版本会在云端应用进程中处理上传文件。请勿上传敏感或受限制数据。"
    )
    assert "云端应用进程" in SIDEBAR_PRIVACY_NOTICE
    assert "敏感或受限制数据" in SIDEBAR_PRIVACY_NOTICE
    detailed_terms = (
        "账号密码",
        "API密钥",
        "交易凭证",
        "个人敏感信息",
        "商业机密",
    )
    for term in detailed_terms:
        assert term not in SIDEBAR_PRIVACY_NOTICE


@pytest.mark.parametrize("page_name", ("首页", "使用说明"))
def test_main_privacy_pages_keep_complete_sensitive_data_warning(
    page_name: str,
) -> None:
    app = _load_app()
    if page_name != "首页":
        _open_page(app, page_name)
    main_text = _visible_text(app.main)

    assert PUBLIC_PRIVACY_NOTICE in main_text
    for term in SENSITIVE_WARNING_TERMS:
        assert term in main_text


@pytest.mark.parametrize(
    "page_name",
    ("首页", "单实验分析", "多实验比较", "参考文件", "使用说明"),
)
def test_all_pages_display_version_from_shared_config(page_name: str) -> None:
    app = _load_app()
    if page_name != "首页":
        _open_page(app, page_name)

    page_text = _visible_text(app)

    assert APP_VERSION == "0.2.0"
    assert "v0.2.0" in page_text
    assert "v0.1.0-rc1" not in page_text
    assert "V0.1.0-RC1" not in page_text
    assert f"V{APP_VERSION.upper()}" not in page_text


def test_generic_import_uses_public_workflow_names_without_stage_codes() -> None:
    app = _ready_generic_return()
    page_text = _visible_text(app)
    button_labels = [button.label for button in app.button]

    assert not app.exception
    assert "字段识别建议" in page_text
    assert "确认字段映射" in page_text
    assert "标准化转换与数据质量预检" in page_text
    assert "严格协议验证" in page_text
    assert all(code not in page_text for code in ("B.2", "B.3", "B.4A", "B.4B"))
    assert button_labels.count("确认字段映射") == 1
    assert button_labels.count("生成标准化预览") == 1
    assert button_labels.count("执行严格协议验证") == 1


def test_quick_guide_version_keeps_configured_case() -> None:
    app = _open_page(_load_app(), "使用说明")
    kicker_markup = [item.value for item in app.markdown if 'class="qrw-kicker"' in str(item.value)]
    style_markup = next(item.value for item in app.markdown if "<style>" in str(item.value))

    assert kicker_markup == [f'<p class="qrw-kicker">快速指南 · v{APP_VERSION}</p>']
    assert "text-transform: none;" in style_markup
    assert "text-transform: uppercase;" not in style_markup


def test_ui_modules_do_not_hardcode_release_version() -> None:
    ui_paths = (
        Path("app.py"),
        Path("src/ui_common.py"),
        Path("src/ui_single.py"),
        Path("src/ui_comparison.py"),
        Path("src/ui_reference_files.py"),
    )

    for path in ui_paths:
        source = path.read_text(encoding="utf-8")
        assert "0.1.0-rc1" not in source
        assert "0.2.0" not in source
    assert all("APP_VERSION.upper()" not in path.read_text(encoding="utf-8") for path in ui_paths)


@pytest.mark.parametrize(
    "page_name",
    ("首页", "单实验分析", "多实验比较", "参考文件", "使用说明"),
)
def test_all_navigation_pages_have_no_uncaught_exception(page_name: str) -> None:
    app = _load_app()
    if page_name != "首页":
        _open_page(app, page_name)

    assert not app.exception
